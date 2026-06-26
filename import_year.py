#!/usr/bin/env python3
"""
Generic importer: Honduras municipal financial data → Supabase.
Source: Base_Financiera_Municipal_2002_2025.xlsx (one sheet per year).

Usage:
    python import_year.py --year 2022
    python import_year.py --year 2021

Behavior:
    1. Deletes ALL existing rows for the target year from Supabase.
    2. Reads 298 rows from the matching sheet in the master Excel.
    3. Inserts them in batches of 50.
    4. Verifies final count and spot-checks key municipalities.

Requirements:
    pip install openpyxl supabase

.env.local must contain:
    REACT_APP_SUPABASE_URL=...
    SUPABASE_SERVICE_ROLE_KEY=...

Execution log:
    2026-06-25 --year 2022 — 298 filas reemplazadas, 0 errores. Fuente: Base_Financiera_Municipal_2002_2025.xlsx
                             Totales: recaudado L 20.93B, propios L 11.46B (ref SEFIN: L 20.9B / L 11.5B). OK.
                             Nota: ingresos_propios era NULL en los 298 registros previos — re-import desde fuente oficial.
                             Nota: DC aparece como "Tegucigalpa, D.C." en hoja 2022 (nombre historico SEFIN).
    2026-06-25 --year 2021 — 298 filas reemplazadas, 0 errores. Fuente: Base_Financiera_Municipal_2002_2025.xlsx
                             Totales: recaudado L 21.42B, propios L  9.46B (ref SEFIN: L 21.4B / L  9.6B). OK.
                             Nota: 27 municipios tenian ingresos_recaudados/propios NULL — reemplazado conjunto completo.
"""

import argparse
import os
import sys
from pathlib import Path
from collections import Counter

# ── Args ──────────────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser(description="Import one year of SEFIN municipal data into Supabase.")
parser.add_argument("--year", type=int, required=True, help="Fiscal year to import (e.g. 2022)")
args_cli = parser.parse_args()
YEAR = args_cli.year

# ── Load .env.local ───────────────────────────────────────────────────────────
def load_env(path):
    p = Path(path)
    if not p.exists():
        return
    with open(p, encoding="utf-8-sig") as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                key, _, val = line.partition("=")
                os.environ.setdefault(key.strip(), val.strip())

load_env(Path(__file__).parent / ".env.local")

# ── Dependency checks ─────────────────────────────────────────────────────────
try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl no instalado. Ejecuta: pip install openpyxl")
    sys.exit(1)

try:
    from supabase import create_client
except ImportError:
    print("ERROR: supabase no instalado. Ejecuta: pip install supabase")
    sys.exit(1)

# ── Config ────────────────────────────────────────────────────────────────────
EXCEL_FILE  = Path(__file__).parent / "Base_Financiera_Municipal_2002_2025.xlsx"
SHEET_NAME  = str(YEAR)
DATA_START  = 7
BATCH_SIZE  = 50
EXPECTED    = 298

SUPABASE_URL     = os.environ.get("REACT_APP_SUPABASE_URL", "")
SERVICE_ROLE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")

# ── Pre-flight ────────────────────────────────────────────────────────────────
if not SUPABASE_URL:
    print("ERROR: REACT_APP_SUPABASE_URL no encontrado en .env.local")
    sys.exit(1)
if not SERVICE_ROLE_KEY:
    print("ERROR: SUPABASE_SERVICE_ROLE_KEY no encontrado en .env.local")
    sys.exit(1)
if not EXCEL_FILE.exists():
    print(f"ERROR: Archivo no encontrado: {EXCEL_FILE}")
    sys.exit(1)

# ── Column map (1-indexed Excel col → db_column) ─────────────────────────────
COL_MAP = {
    3:  ("code",                        "int"),
    4:  ("department",                  "str"),
    5:  ("name",                        "str"),
    7:  ("population",                  "int"),
    9:  ("presupuesto_municipal",       "float"),
    10: ("gastos_presupuestados",       "float"),
    11: ("ingresos_propios",            "float"),
    12: ("ingresos_recaudados",         "float"),
    13: ("autonomia_financiera",        "float"),
    14: ("ingresos_corrientes",         "float"),
    16: ("ingresos_tributarios",        "float"),
    17: ("impuesto_bi",                 "float"),
    18: ("impuesto_personal",           "float"),
    19: ("impuesto_industria",          "float"),
    20: ("impuesto_comercio",           "float"),
    21: ("impuesto_servicios",          "float"),
    22: ("impuesto_pecuario",           "float"),
    23: ("impuesto_extraccion",         "float"),
    24: ("impuesto_telecomunicaciones", "float"),
    25: ("tasas_servicios",             "float"),
    26: ("derechos",                    "float"),
    27: ("ingresos_no_tributarios",     "float"),
    29: ("ingresos_capital",            "float"),
    30: ("prestamos",                   "float"),
    31: ("venta_activos",               "float"),
    32: ("contribuciones",              "float"),
    33: ("colocacion_bonos",            "float"),
    34: ("transferencias_art91",        "float"),
    35: ("otras_transferencias",        "float"),
    36: ("subsidios",                   "float"),
    37: ("herencias_legados",           "float"),
    38: ("otros_ingresos_capital",      "float"),
    39: ("recursos_balance",            "float"),
    41: ("gastos_funcionamiento",       "float"),
    42: ("servicios_personales",        "float"),
    43: ("servicios_no_personales",     "float"),
    44: ("materiales_suministro",       "float"),
    45: ("transferencias_corrientes",   "float"),
    46: ("otros_gastos",                "float"),
    48: ("gastos_capital_deuda",        "float"),
    49: ("bienes_capitalizables",       "float"),
    50: ("transferencias_capital",      "float"),
    51: ("activos_financieros",         "float"),
    52: ("servicios_deuda",             "float"),
    53: ("otros_gastos_capital",        "float"),
    54: ("asignaciones_globales",       "float"),
    56: ("total_egresos",               "float"),
    57: ("superavit_deficit",           "float"),
    58: ("gasto_corriente",             "float"),
    59: ("ingreso_corriente_ajustado",  "float"),
}

# ── Helpers ───────────────────────────────────────────────────────────────────
def cast(value, dtype):
    if value is None or str(value).strip() == "":
        return None
    try:
        if dtype == "int":   return int(float(value))
        if dtype == "float": return float(value)
        return str(value).strip()
    except (ValueError, TypeError):
        return None

def is_totals_row(row):
    v = row[0]
    if v is None:
        return True
    try:
        int(float(v))
        return False
    except (ValueError, TypeError):
        return True

# ── Connect ───────────────────────────────────────────────────────────────────
print(f"Importando year={YEAR} — conectando a Supabase...")
client = create_client(SUPABASE_URL, SERVICE_ROLE_KEY)

# ── Step 1: delete existing rows for this year ────────────────────────────────
existing = client.table("municipalities").select("year", count="exact").eq("year", YEAR).execute().count
print(f"\nStep 1 — Eliminando {existing} filas existentes para year={YEAR}...")
if existing > 0:
    client.table("municipalities").delete().eq("year", YEAR).execute()
    after_delete = client.table("municipalities").select("year", count="exact").eq("year", YEAR).execute().count
    print(f"  Eliminadas. Quedan: {after_delete}")
else:
    print("  Sin filas previas — nada que eliminar.")

# ── Step 2: read Excel ────────────────────────────────────────────────────────
print(f"\nStep 2 — Leyendo {EXCEL_FILE.name} (hoja '{SHEET_NAME}')...")
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

if SHEET_NAME not in wb.sheetnames:
    print(f"ERROR: Hoja '{SHEET_NAME}' no encontrada. Hojas disponibles: {wb.sheetnames}")
    sys.exit(1)

ws = wb[SHEET_NAME]
rows_to_insert = []
skipped = 0

for row in ws.iter_rows(min_row=DATA_START, max_row=DATA_START + EXPECTED, values_only=True):
    if is_totals_row(row):
        skipped += 1
        continue
    record = {"year": YEAR}
    for col_1idx, (db_field, dtype) in COL_MAP.items():
        raw = row[col_1idx - 1]
        if db_field == "department":
            record[db_field] = str(raw).strip().upper() if raw is not None else None
        elif db_field == "name":
            record[db_field] = str(raw).strip().title() if raw is not None else None
        else:
            record[db_field] = cast(raw, dtype)
    rows_to_insert.append(record)

print(f"  Filas a insertar : {len(rows_to_insert)}")
print(f"  Filas omitidas   : {skipped}  (TOTALES / vacías)")

if len(rows_to_insert) != EXPECTED:
    print(f"  ERROR: Se esperaban {EXPECTED} filas, se encontraron {len(rows_to_insert)}. Abortando.")
    sys.exit(1)

# ── Step 3: insert in batches ─────────────────────────────────────────────────
print(f"\nStep 3 — Insertando {len(rows_to_insert)} filas...")
errors = []
inserted = 0

for i in range(0, len(rows_to_insert), BATCH_SIZE):
    batch = rows_to_insert[i : i + BATCH_SIZE]
    try:
        client.table("municipalities").insert(batch).execute()
        inserted += len(batch)
        print(f"  {inserted}/{len(rows_to_insert)} filas insertadas...")
    except Exception as e:
        err_msg = f"Lote {i}–{i + len(batch) - 1}: {e}"
        errors.append(err_msg)
        print(f"  ERROR — {err_msg}")

print(f"\n{'=' * 55}")
print(f"RESULTADO: {inserted} filas insertadas, {len(errors)} errores")
if errors:
    print("\nDetalle de errores:")
    for err in errors:
        print(f"  • {err}")

# ── Step 4: verify ────────────────────────────────────────────────────────────
print(f"\nStep 4 — Verificación final para year={YEAR}:")
resp = client.table("municipalities").select(
    "name,ingresos_recaudados,ingresos_propios,autonomia_financiera"
).eq("year", YEAR).execute()
rows = resp.data

final_count = client.table("municipalities").select("year", count="exact").eq("year", YEAR).execute().count
total_rec  = sum(r["ingresos_recaudados"] or 0 for r in rows) / 1e9
total_prop = sum(r["ingresos_propios"]    or 0 for r in rows) / 1e9
valid_aut  = [r["autonomia_financiera"] for r in rows if r["autonomia_financiera"] is not None]
avg_aut    = sum(valid_aut) / len(valid_aut) if valid_aut else 0

count_marker = "OK" if final_count == EXPECTED else "REVISAR"
print(f"  Filas:           {final_count} / {EXPECTED}  {count_marker}")
print(f"  Recaudado total: L {total_rec:.2f}B")
print(f"  Propios total:   L {total_prop:.2f}B")
print(f"  AutoFin promedio: {avg_aut:.1f}%")

for anchor in ["Distrito Central", "San Pedro Sula"]:
    hit = next((r for r in rows if r["name"] == anchor), None)
    if hit:
        rec  = hit["ingresos_recaudados"]
        prop = hit["ingresos_propios"]
        aut  = hit["autonomia_financiera"]
        print(f"  {anchor}: recaudado=L{rec/1e6:.1f}M  propios=L{prop/1e6:.1f}M  AutoFin={aut:.1f}%")
    else:
        print(f"  {anchor}: NO ENCONTRADO")
